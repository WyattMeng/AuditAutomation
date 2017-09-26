# -*- coding: utf-8 -*-
"""
Created on Tue Sep 19 14:34:33 2017

@author: Maddox.Meng
"""

'''之前用xlrd、pandas读取，这次尝试用单一模块openpyxl搞定一切

'''

from openpyxl import load_workbook
import re

from openpyxl.utils import get_column_letter, column_index_from_string# 根据列的数字返回字母
#column_index_from_string('aa') == 27

'''正则表达式match日期'''
def matchDate(date):
    #print 'matchDate'    
    # 将正则表达式编译成Pattern对象
    pattern = re.compile(r'2\d\d\d年')     
    # 使用Pattern匹配文本，获得匹配结果，无法匹配时将返回None
    match = pattern.match(date.encode('utf-8')) #u'2017\u5e7412\u670831\u65e5'     
    if match:
        # 使用Match获得分组信息
        #print match.group()
        return True
    if not match:
        return False


'''get一级条目的单元格value list，rownumber list。进而从底稿get每个一级条目的高度。'''
#openpyxl worksheet[x][y], x从1开始计算，y从0开始计算
def getCatalogBands(worksheet, catalogColNo):
    print 'getCatalogBands'
    rownumbers=[]
    sheetnames=[]
    y = catalogColNo                      #假定已知第0列是存放底稿一级条目（GET sheetname）的列
    for x in range(1, worksheet.max_row+1): #openpyxl的行从1起计，所以这里range从1开始。
        if worksheet[x][y].value != None: #append非空单元格value、rownumber
            rownumbers.append(x)
            sheetnames.append(worksheet[x][y].value)
            
    #rownumbers.append(worksheet.max_row)  #最后append上max_row，为了下面dict好计算。因为最后一行一级条目的最大行好就是整个worksheet的最大行号。

    xCatalogBands = {}    #底稿其中一个一级条目的value（也是GET的sheetname）、起始rownumber、结束rownumber字典    
    k=0
    for sheetname in sheetnames:
        if k == len(rownumbers) - 1:
            x_max = worksheet.max_row 
        else:
            x_max = rownumbers[k+1] - 1
        xCatalogBands[k] = {'sheetname':sheetname, 'x_min':rownumbers[k], 'x_max':x_max}
        k+=1
    return xCatalogBands   


def getXtitles(xCatalogBands):
    print 'getXtitles'
    '''得到xtitles，包括title name和title行号'''    
    y=2  #对于一个一级科目（一个GET sheet）的Area，判断第3 or 第4列是写有二三级科目的列
    xtitles_a = []
    #print 'Below is [' + xdict[k]['sheetname'] + ']:'
    for x in range(xCatalogBands[k]['x_min'], xCatalogBands[k]['x_max'] + 1):        
        if (
            ws[x][y].value != None and 
            ws[x][y].value != u'项目' and
            ws[x][y].value != u'附注六' and
            ws[x][y].value != u'股东名称' and
            ws[x][y].value != u'关联方名称' and
            matchDate(ws[x][y].value) is False
            ):#去掉项目、附注六、股东名称、关联方名称、年份, yitems行数 = 二级科目行 - 一级科目行 - 1
            xitem = ws[x][y].value
            x     = ws[x][y].row
            xtitles_a.append([xitem, x])
            
    y=3  #对第4列循环
    xtitles_b = []
    #print 'Below is [' + xdict[k]['sheetname'] + ']:'
    for x in range(xCatalogBands[k]['x_min'], xCatalogBands[k]['x_max'] + 1):
        #正则表达式去掉2xxx年的匹配cell
        if (
            ws[x][y].value != None and 
            isinstance(ws[x][y].value,long) is False and 
            isinstance(ws[x][y].value,float) is False and
            matchDate(ws[x][y].value) is False
            ):
            xitem = ws[x][y].value
            x     = ws[x][y].row
            xtitles_b.append([xitem, x])

    if len(xtitles_a) <= len(xtitles_b):  #如果列3非空个数 小于等于 列4非空非数字个数，则列4为title列
        xtitles = xtitles_b
    else:                                 #反之，列2为title列
        xtitles = xtitles_a
        
    return xtitles    #[['资产', 5],['负债', 8],['合计', 9],...]


def getYRows(xtitles, xCatalogBands):
    print 'FUNCTION [getYRows]'
    
    '''得到ytitles行数（在后面起到作用）'''
    yrows = xtitles[0][1] - xCatalogBands[k]['x_min'] - 1  #根据excel模板规律，ytitle行数 = 第1个xtitle行号 - 所在catalog行号 - 1
    
    '''得到ytitles，包括title name和title列号'''
    ytitleRowlist = []
    for i in range(0, yrows):
        ytitleRowlist.append(xCatalogBands[k]['x_min'] + i)
    return ytitleRowlist           #like [5, 6] or [92]



def getYColumnNo(ytitleRowlist):
    print 'getYColumnNo'
    ytitleCollist = []
    for x in ytitleRowlist:
        for y in range(4, ws.max_column): #ytitle都在第5、6列起始。特例：R利率敏感性分析，没有ytitle。
            #1.对yrowslist遍历
            #2.找到非空单元格，记录到list
            if ws[x][y].value != None:
                #print ws[x][y].value, ws[x][y].column
                ytitleCollist.append(ws[x][y].column)
                ytitleCollist = sorted(list(set(ytitleCollist))) #去重、排序
    return ytitleCollist            
            

def getMergedCells(worksheet): #得到底稿里一个sheet的所有合并格，然后确定某个合并格属于哪个catalog
    print 'getMergedCells'
    #mergedCellsList = worksheet.merged_cell_ranges #['E81:K81','N81:T81','E5:H5',...]
                  #worksheet.merged_cell {'E39','E5','E81','F39',...}
    mergedCellsList = []  #Debug:一开始这里没设定，导致list在整个循环不断增长， 最后报错Cell range K5:N5 not known as merged.           
    for mergedCells in worksheet.merged_cell_ranges: 
        #int(merged_cells.split(':')[0][1:]) 
        if int(re.findall("\d+",mergedCells)[0]) in range(xCatalogBands[k]['x_min'], xCatalogBands[k]['x_max'] + 1):   
            #print 'mergedCells',mergedCells
            mergedCellsList.append(mergedCells)        
        
    return mergedCellsList                  


def unmergeCells(worksheet, mergedCellsList):
    for mergedCells in mergedCellsList:
        worksheet.unmerge_cells(mergedCells)
  
    
#def getRelevantMergedCells(mergedCellsList): #当前Catalog里merged cell
#    relevantMergedCellsList = []
#    for mergedCells in mergedCellsList:        
#        #int(merged_cells.split(':')[0][1:]) 
#        if int(re.findall("\d+",mergedCells)[0]) in range(xCatalogBands[k]['x_min'], xCatalogBands[k]['x_max']):   
#            print 'mergedCells',mergedCells
#            relevantMergedCellsList.append(mergedCells)
#    return relevantMergedCellsList   


'''拆分合并格，并在有数据的列写入合并格的值（一切动作在内存中完成）'''
def splitMergedCellValue(worksheet, mergedCellsList, ytitleCollist):
    print 'splitMergedCellValue'
    print ytitleCollist
    for mergedCells in mergedCellsList:
        worksheet.unmerge_cells(mergedCells)
        for cells in ws[mergedCells]:
            #print 'cells', cells
            for cell in cells:
                #print 'column',cell.column
                if ytitleCollist.count(cell.column) > 0 :
                    #print mergedCells.split(':')[0]
                    cell.value = ws[mergedCells.split(':')[0]].value
                      
'''ytitle'''        
def getYtitles(ytitleCollist):
    print 'getYtitles'
    ytitles = []
    for y in ytitleCollist:
        ySingleTitle = []
        for x in ytitleRowlist:
            coor = y+str(x)
            #print ws[coor].value
            ySingleTitle.append(ws[coor].value)
        ySingleTitle.append(y)
        ytitles.append(ySingleTitle)    
    return ytitles


'''ytitle2'''        
def getYtitles2(ytitleCollist):
    print 'getYtitles2'
    ytitles = []
    for y in ytitleCollist:
        
        ytitleName = ''
        for x in ytitleRowlist:
            coor = y+str(x)
            if ws[coor].value != None:
                ytitleName = ytitleName + ws[coor].value + '>>>'
        print ytitleName.strip('>>>')
        ytitles.append([ytitleName.strip('>>>'),y])
    #print ytitles
    return ytitles    



'''生成最终的dict'''
import json
def createDataDict(xtitles, ytitles):  
    print 'createDataDict'
    dict={}
    k=0    
    for ytitle in ytitles:
        #print ytitle
        for xtitle in xtitles:  
            coor = ytitle[-1]+str(xtitle[-1])
            #print 'coor = ',coor
            if ws[coor].value != None: #除掉了二级科目，因为他们没有对应数值   
                #print 'coor = ',coor
                dict[k]={'xtitle':xtitle[0], 'ytitle':ytitle[0], 'value':ws[coor].value}
                k+=1

    #打印中文字典
    print json.dumps(dict, encoding='UTF-8', ensure_ascii=False)                           
    return dict  


'''根据数据源dict写入GET'''

def getGETArea(ytitleCollist):
    max_col = get_column_letter(column_index_from_string(ytitleCollist[-1]) - 1)
    max_row = xtitles[-1][-1] - xCatalogBands[k]['x_min'] +1 +1
    coor = max_col + str(max_row)
    return coor
    

def xxx(wb_dest, coor, xCatalogBands, dict): #注意有两个dict，此k非彼k
    listx=[]
    #listy=[]
    wsx = wb_dest.get_sheet_by_name(xCatalogBands[k]['sheetname'])   
    for key in dict:         
        for col in wsx['A1:%s' % coor]:
    
            for cell in col:
    
                if cell.value == dict[key]['xtitle']:
                    listx.append([key, cell.row, dict[key]['value']])
                    
    return listx     #[[0, 5, 1324L], [1, 5, 9688L]]           


#对底稿处理ytitle的过程移植到GET么，很麻烦的感觉，开始吧。想想从哪儿开始。
def yyy(wb_dest, coor, xCatalogBands, dict):
    listy=[]
    wsx = wb_dest.get_sheet_by_name(xCatalogBands[k]['sheetname'])
    for key in dict:         
        for col in wsx['A1:%s' % coor]:    
    
            for cell in col:

                if isinstance(cell.value, str) is True and cell.value.find('=') !=-1:
                    datesheet = cell.value.replace('=','').split('!')[0] #Instraction
                    datecell  = cell.value.replace('=','').split('!')[1] #B1 B5
                    datevalue = wb_dest[datesheet][datecell].value       #2017年12月31日
                    if datevalue == dict[key]['ytitle']:
                        listy.append([key, cell.column, dict[key]['value']])                
                

'''----------------------------------------------------------------------------------------------------'''
path = r'C:\Workspace\测试\risk table.xlsx'.decode('utf-8') #r和decode utf-8组合使用可以解决中文路径的issue
wb = load_workbook(path)
#ws = wb.get_sheet_names() #[u'BS&PL', u'CF']
ws = wb.get_sheet_by_name(u'信用风险')
xCatalogBands = getCatalogBands(ws,0)

path_dest = r'C:\Workspace\测试\GET.xlsx'.decode('utf-8')
wb_dest = load_workbook(path_dest)

for k in xCatalogBands:
    print 'k',k
    
    xtitles = getXtitles(xCatalogBands)
    
    ytitleRowlist = getYRows(xtitles, xCatalogBands)
    ytitleCollist = getYColumnNo(ytitleRowlist)
    
    mergedCellsList = getMergedCells(ws)
    splitMergedCellValue(ws, mergedCellsList, ytitleCollist)
    
    ytitles = getYtitles2(ytitleCollist)
    dataDict = createDataDict(xtitles, ytitles)

    '''根据数据源dict写入GET'''

    coor = getGETArea(ytitleCollist)
    listx = xxx(wb_dest, coor, xCatalogBands, dataDict)
    listy = yyy(wb_dest, coor, xCatalogBands, dataDict)
    
#wb.save(r'C:\Workspace\测试\copy-tst-merged.xlsx'.decode('utf-8'))    
    
    
#    '''yitems'''    
#    ylist =[]
#    for x in yrowslist:
#        for y in range(3, ws.max_column): #预知从列3开始
#            if ws[x][y].value != None:
#                ylist.append(ws[x][y].column)
#                print ws[x][y].coordinate, ws[x][y].value 
#                print ylist
#                print list(set(ylist)).sort()
            
#for item in ws.merged_cell_ranges:
#    print item.split(':')[0]            
    
    
'''----------------------------------------------------------
print json.dumps(dict, encoding='UTF-8', ensure_ascii=False) 


'''   