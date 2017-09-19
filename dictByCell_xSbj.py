# -*- coding: utf-8 -*-
"""
Created on Thu Sep 14 01:55:54 2017

@author: mmmkk
"""

#import pandas as pd
#print pd.__version__
#import xlrd
#print xlrd.__VERSION__
#import openpyxl
#print openpyxl.__version__
import sys
#path = r'C:\Users\mmmkk\Desktop\测试\tst.xlsx'.decode('utf-8')
path = r'C:\Workspace\测试\tst.xlsx'.decode('utf-8')
#print path

from xlrd import open_workbook
import openpyxl
wb = open_workbook(path)
#print wb.sheet_names()
ws = wb.sheet_by_index(0)

#print ws.name, ws.nrows, ws.ncols

rows = ws.row_values(6)
cols = ws.col_values(4)
#print rows

#for i in cols:
#    if i !='':
#        print i,ctype(i)

#l = [u'中文',u'英文',u'Eng']
#msg = repr([x.encode(sys.stdout.encoding) for x in l]).decode('string-escape')
#print msg

# 获取单元格内容
#print ws.cell(4,0).value#.encode('utf-8')
#print ws.cell_value(6,2)#.encode('utf-8')
#print ws.row(6)[4].value#.encode('utf-8')

# 获取单元格内容的数据类型
#print ws.cell(6,4).ctype
#ctype : 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error

#print ws.merged_cells #[(71, 72, 5, 21), (87, 88, 5, 21)]
#print ws.cell_value(71,5)
#print ws.cell_value(87,5)

#for cell in ws.col_values(0):
#    print cell.ctype

def xxx(worksheet):
    xitems=[]
    y=2
    for x in range(0,ws.nrows):
        if ws.cell(x,y).ctype !=0:        
            xitem = ws.cell_value(x,y)
            xitems.append([xitem,x])
            
    yitems=[]
    x=0
    for y in range(0+1,ws.ncols):
        if ws.cell(x,y).ctype !=0 and ws.cell(x,y).value != u'附注六':
            yitem = ws.cell_value(x,y)
            yitems.append([yitem,y])        
    
    dict={}
    i=0
#    for [xitem,x] in xitems:
#        for [yitem,y] in yitems:
    for [yitem,y] in yitems:
        for [xitem,x] in xitems:            
            if ws.cell([xitem,x][1], [yitem,y][1]).ctype !=0: #除掉了二级科目，因为他们没有对应数值
                #print [xitem,x][0], [yitem,y][0], ws.cell_value([xitem,x][1], [yitem,y][1])
                dict[i]={'xitem':[xitem,x][0], 'yitem':[yitem,y][0], 'value':ws.cell_value([xitem,x][1], [yitem,y][1])}
                i+=1
                
    return dict            


'''根据数据源dict写入GET'''
dict = xxx(ws)
#pathd = r'C:\Users\mmmkk\Desktop\测试\GET.xlsx'.decode('utf-8')
#pathd = r'C:\Workspace\测试\GET.xlsx'.decode('utf-8')
pathd = r'C:\Workspace\测试\Grace-Excel Template (b)_Non-listed_170816-哈尔滨村镇银行8-18-GTH.xlsx'.decode('utf-8')


wbx = openpyxl.load_workbook(pathd)
wsx = wbx.get_sheet_by_name(u'BS_资产')
#print dict


listx=[]
listy=[]
for k in dict:
#    print k,dict[k]['xitem'],dict[k]['yitem'],dict[k]['value']
    for col in wsx['A1:G100']:

        for cell in col:

            if cell.value == dict[k]['xitem']:
                listx.append([k,cell.row,dict[k]['value']])
                dict[k]['yitem']

#                print k,cell.value,cell.row,cell.column 
            
            elif isinstance(cell.value,str) is True and cell.value.find('=') !=-1:
                datesheet = cell.value.replace('=','').split('!')[0]
                datecell  = cell.value.replace('=','').split('!')[1]
                datevalue = wbx[datesheet][datecell].value
                if datevalue == dict[k]['yitem']:
                    listy.append([k,cell.column,dict[k]['value']])
#                    print k,cell.value,cell.row,cell.column
#    wsx.cell(row=xbase,column=ybase).value = cell

cells4Fill={}
i=0
for itemx in listx:
    for itemy in listy:
        if itemx[0] == itemy[0]:
            cell = itemy[1] + str(itemx[1])
            value = itemx[2]
            cells4Fill[i] = {'cell': cell, 'value': value}
            i+=1    
print cells4Fill    
#print wsx['D2'].value
#print wbx['Instraction']['B1'].value
for k in cells4Fill:
    wsx[cells4Fill[k]['cell']] = cells4Fill[k]['value']

wbx.save(pathd)




#list二级subject name
def list2ndSbj(worksheet):
    y=2    
    xitems=[]
    for x in range(0,ws.nrows):
        if ws.cell(x,y).ctype !=0:        
            xitem = ws.cell_value(x,y)
            xitems.append([xitem,x])
            
    yitems=[]
    for y in range(1,ws.ncols):
        if ws.cell(0,y).ctype !=0 and ws.cell_value(0,y) != u'附注六':
            yitem = ws.cell_value(0,y)
            yitems.append([yitem,y])
    return xitems, yitems        

xitems = list2ndSbj(ws)[0]
yitems = list2ndSbj(ws)[1]

i=0
for [xitem,x] in xitems:
    for [yitem,y] in yitems:
        if ws.cell([xitem,x][1], [yitem,y][1]).ctype ==0:
            pass#print [xitem,x][1],[xitem,x][0], [yitem,y][0], ws.cell_value([xitem,x][1], [yitem,y][1])
            #dict[i]={'xitem':[xitem,x][0], 'yitem':[yitem,y][0], 'value':ws.cell_value([xitem,x][1], [yitem,y][1])}
            i+=1
            
aSbj = u'BS_资产'           
bSbj = ws.cell_value(2,2)
y=2
for x in range(2+1,14):
    cSbj = ws.cell_value(x,y)
#    print aSbj+'>>>'+bSbj+'>>>'+cSbj