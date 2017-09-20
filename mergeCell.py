# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import xlrd
print xlrd.__VERSION__ #1.0.0 >>> 1.1.0
import openpyxl
print openpyxl.__version__ #2.4.7 >>> 2.5.0a3
import pandas as pd
print pd.__version__ #0.20.1 >>> 0.20.3
import re
print re.__version__

    
    
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string# 根据列的数字返回字母  

path = r'C:\Workspace\测试\tst-merged.xlsx'.decode('utf-8')
wb = load_workbook(path)
ws = wb.get_sheet_by_name('Sheet1') 

print ws.merged_cell_ranges
 
#wb.save(path)
print ws.merged_cell_ranges

ws.unmerge_cells('D1:J1')
ylist = ['A','D','G','J','M']
#print ws['D1:J1']
for cells in ws['D1:J1']:
    for cell in cells:
        if ylist.count(cell.column) > 0 :
            cell.value = u'已逾期未减值'
            print cell.column, cell.value
#wb.save(r'C:\Workspace\测试\copy-tst-merged.xlsx'.decode('utf-8'))        
yitem = ''
for y in ylist:
    for x in range(1,3):
        coor = y+str(x)
#        print y,x,ws[coor].value
    if ws[coor].value != None:
        yitem = yitem + ws[coor].value
    print yitem    
'''ws['A0']代表第一列'''

k=0
ydict = {}
for y in ylist:
    yitem = ws[]
    ydict[k] = y
    k+=1

        


