# -*- coding: utf-8 -*-
"""
Created on Sat Sep 09 17:52:59 2017

@author: Maddox.Meng
"""

import win32ui
import os
import timeit
 
dlg= win32ui.CreateFileDialog(1)# 1表示打开文件对话框
dlg.SetOFNInitialDir('%USERPROFILE%\desktop')# 设置打开文件对话框中的初始显示目录
dlg.DoModal()
 
filename= dlg.GetPathName()# 获取选择的文件名称
parent_path = os.path.dirname(filename)
#print parent_path.decode('gbk')
path2=parent_path
'''----------------------------------------------------------'''
dlg= win32ui.CreateFileDialog(1)# 1表示打开文件对话框
dlg.SetOFNInitialDir('%USERPROFILE%\desktop')# 设置打开文件对话框中的初始显示目录
dlg.DoModal()
 
filename2= dlg.GetPathName()# 获取选择的文件名称

#print filename2.decode('gbk')
dest=filename2


'''=============================================================================='''



#print('parent_path = %s' % parent_path) 
#file_name = os.path.split(file_path)[-1] 
#print('file_name = %s' % file_name)

import pandas as pd
import numpy as np
from xlrd import open_workbook
from numpy import float64

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
#import os.path
s = os.sep #根据unix或win，s为\或/
#path = r'C:\Users\maddox.meng\Desktop\哈尔滨GET表格及贷款、存款WP和贷款PBC\底稿'
#root = 'c:'+s+'Users\maddox.meng\Desktop\哈尔滨GET表格及贷款、存款WP和贷款PBC\底稿'+s
#path2= 'C:\\Users\\maddox.meng\\Desktop\\automation'
#dest = 'C:\\Users\\maddox.meng\\Desktop\\GET.xlsx'




'''1'''#list文件夹里所有底稿文件 
awps = []
for i in os.listdir(path2):
    if os.path.isfile(os.path.join(path2,i)):
        awps.append(os.path.join(path2,i))
        
'''4'''
#确定某excel的某sheet的第几列存放对应GET表的sheetname。逻辑是出现文字的第一列
#这个excel的第0列是sheetname
def findSheetnameCol(dataframe):
    for y in range(0,df.shape[-1]):
        x = 0
        for cell in df[y]:
            if cell is not np.nan:
                #print x,y,cell
                #break
                return y
            x+=1   
            
            
'''5'''
#列出底稿里某sheet的某column里存放的sheetname和row number
def listSheetNames(col):
    sheetnames=[]
    rownumbers=[]
    x = 0       
    for cell in df[col]:
        if cell is not np.nan:
            #print x,cell
            sheetnames.append(cell)
            rownumbers.append(x)
        x+=1    
    return sheetnames,rownumbers 


'''6'''
#根据上面的row number，加上df.shape可以确定每个sheetname的area
#[4, 41, 71, 87], df.shape = (101, 21), [4:41,0:21],[41:71,0:21],[71:87,0:21],[87:101,0:21]
def getDictReportArea(sheetnames,rownumbers):
    dictReortAreas={}
    for i in range(0,len(sheetnames)):
        if i < len(sheetnames) - 1 :
    #        x_bot = rownumbers[i+1]
    #        area = '[%s:%s, 0:%s]' % (rownumbers[i],x_bot,df.shape[1])
    #        dict[i] = ({'sheet':sheetnames[i],
    #                   'area': area})
            x_min=rownumbers[i]
            x_max=rownumbers[i+1]
            y_min=0
            y_max=df.shape[1]
            
            dfReportArea = df.iloc[x_min:x_max, y_min:y_max]
            #dictReortAreas[i] = {'sheet':sheetnames[i], 'x_min':rownumbers[i], 'x_max':rownumbers[i+1], 'y_min':0,'y_max':df.shape[1]}
            dictReortAreas[i] = {'sheet':sheetnames[i], 'dfReportArea':dfReportArea}
        else:
    #        x_bot = df.shape[0]
    #        area = '[%s:%s, 0:%s]' % (rownumbers[i],x_bot,df.shape[1])
    #        dict[i] = ({'sheet':sheetnames[i],
    #                   'area': area})
            x_min=rownumbers[i]
            x_max=df.shape[0]
            y_min=0
            y_max=df.shape[1]
            
            dfReportArea = df.iloc[x_min:x_max, y_min:y_max]    
            #dictReortAreas[i] = {'sheet':sheetnames[i], 'x_min':rownumbers[i], 'x_max':df.shape[0], 'y_min':0,'y_max':df.shape[1]}
            dictReortAreas[i] = {'sheet':sheetnames[i], 'dfReportArea':dfReportArea}
    
    return dictReortAreas #存储了某个10000表里所有sheetname和对应的dataArea
    
            


'''7'''
#dataArea
'''dataArea的高度--方法2'''
#得到BS_资产的data area是
def getReportDataArea(dfReportArea):
    
    xList=[]
    for y in range(2,4): #从第二行开始，对第2、3列循环
        #第3列空单元格是numpy.float64
        x=1
        for cell in dfReportArea.iloc[1:,y]:
            if cell is not np.nan and isinstance(cell,float64) is False:
                #print cell,x,y
                #print x,cell
                xList.append(x) #xList[0], xList[-1]
            x+=1
    #print 'max',max(xList), min(xList)
    
    '''dataArea的宽度--方法2'''
    #日期在dfArea的第一行
    yList=[]
    for x in range(0,2): 
        y=4
        for cell in dfReportArea.iloc[x,4:]:
            #if isinstance(cell,unicode) is True and cell.find('201') !=-1:
            if cell is not np.nan and isinstance(cell,float64) is False and cell.find(u'附注六') == -1:    
                #print cell, x, y
                yList.append(y)
            y+=1
   
    x_min=min(xList)
    x_max=max(xList)
    y_min=min(yList)
    y_max=max(yList)
     
    dfReportDataArea = dfReportArea.iloc[x_min:x_max+1, y_min:y_max+1]  
    return dfReportDataArea       

def getReportDataArea2(dfReportArea):
    x_min=2
    x_max=3
    y_min=4
    y_max=7
    dfReportDataArea = dfReportArea.iloc[x_min:x_max+1, y_min:y_max+1]  
    return dfReportDataArea


def loadWorkbook(dest_xlsx):
    wb = load_workbook(filename = dest_xlsx, data_only = False)
    return wb

'''8''' 
#目标xlsx GET表被插入位置的起始点          
'''---------------------panda会把年份公式读成float或float64--------------------'''
'''---------------------openpyxl会把年份公式读成str，空为NoneType--------------------'''
def getBasepointDest(wb, sheetname):
    '''openpyxl写入GET。参见modifyExcel.py'''
    #wb = load_workbook(filename = dest_xlsx, data_only = False)
    #with load_workbook(filename = dest_xlsx, data_only = False) as wb:
    ws = wb.get_sheet_by_name(sheetname)
    
    '''用openpyxl计算插入dest数据的起始坐标x'''
    xList=[]
    for y in range(1,4): #第1、2列
        
        #openpyxl计算总大小ws.max_row
        for x in range(1, ws.max_row+1):
            cell = ws.cell(row=x,column=y).value
            if (isinstance(cell,unicode) and
                cell.find(u'项目') ==-1 and
                cell.find(u'附注六') ==-1 and
                cell.find(u'股东名称') ==-1 and
                cell.find(u'关联方名称') ==-1            
                ):
                #print x,ws.max_row,cell
                xList.append(x)
                
            
    '''用openpyxl计算插入dest数据的起始坐标y'''
    yList=[]
    for x in range(1,4):
        
        for y in range(3,10):
            cell = ws.cell(row=x,column=y).value
            #print ws.cell(row=x,column=y).value, type(ws.cell(row=x,column=y).value)
            if ((isinstance(cell,str) or isinstance(cell,unicode)) and
                cell.find(u'项目') ==-1 and
                cell.find(u'附注六') ==-1 and
                cell.find(u'股东名称') ==-1 and
                cell.find(u'关联方名称') ==-1            
                ):
                #print y,cell
                yList.append(y)
    #print 'xList= ',xList
    #print 'yList= ',yList               
    return min(xList),min(yList)   


def getBasepointDest2(dest_xlsx, sheetname):
    xList=[4]
    yList=[4]
    return min(xList),min(yList)

'''9'''
#写入GEt表
def writeToDest(wb,sheetname,x,y,dataframe):
    #wb = load_workbook(filename = dest_xlsx, data_only = False)
    #with load_workbook(filename = dest_xlsx, data_only = False) as wb:
    ws = wb.get_sheet_by_name(sheetname)
    
    xbase = x
    for row in dataframe_to_rows(dataframe, index=False, header=False):
        #ws.cell(row=x,column=4).value = r[0]
        
        ybase = y
        for cell in row:
            ws.cell(row=xbase,column=ybase).value = cell           
            ybase+=1
        xbase+=1
    wb.save(dest)         


import sys
reload(sys)
sys.setdefaultencoding('utf8')


from Tkinter import *
#import tkfont

root = Tk()
text = Text(root)
text.configure(font=("微软雅黑", 8))
#font = tkFont.Font(family:'Microsoft-Yahei-UI-Light')
#text.insert(INSERT, "Hello.....")
#text.insert(END, "Bye Bye.....")
#text.pack()



logfile = open('logs.txt', 'w+')
'''2'''
start = timeit.default_timer() 
#遍历每个底稿    
for awp in awps:
    #awp = awps[0]
    print awp
    
    logfile.write('u| -- '+awp.split('\\')[-1]+'\n')
    text.insert(INSERT, '|-- '+awp.split('\\')[-1]+'\n')
    text.update()
    text.see('end')
    
    #'''3'''#遍历一个底稿的每个sheet
    wb = open_workbook(awp) 
    print wb.sheet_names()
    #遍历某个底稿的每个sheet
    for sheet in wb.sheet_names():
        
        logfile.write(u'  | -- '+sheet+'\n')
        text.insert(INSERT, u'  | -- '+sheet+'\n')
        text.update()
        text.see('end')
        
        #print sheet
        df = pd.read_excel(awp, sheet, header=None)
        colWithSheetname = findSheetnameCol(df)
        sheetnames = listSheetNames(colWithSheetname)[0]
        rownumbers = listSheetNames(colWithSheetname)[1]
        dictReportArea = getDictReportArea(sheetnames,rownumbers)
        
        #遍历每个report item
        for key in dictReportArea:
            sheetname = dictReportArea[key]['sheet'] #report items, GET's sheetnames
            dfDataArea = dictReportArea[key]['sheet']
            
            if dictReportArea[key]['sheet'].find(u'利率敏感性分析') !=-1:
                dfReportDataArea = getReportDataArea2(dictReportArea[key]['dfReportArea'])
            else:    
                dfReportDataArea = getReportDataArea(dictReportArea[key]['dfReportArea']) 
            #dfReportDataArea = getReportDataArea(dictReportArea[key]['dfReportArea'])
            #print dfReportDataArea
            #print 'sheetname= ',sheetname
            
            logfile.write(u'    | -- '+sheetname+'\n')
            text.insert(INSERT, '    | -- '+sheetname+'\n')
            text.update()
            text.see('end')
            text.pack()
            
            wb = loadWorkbook(dest)
            
            if dictReportArea[key]['sheet'].find(u'利率敏感性分析') !=-1:
                xBase = getBasepointDest2(wb, sheetname)[0]
                yBase = getBasepointDest2(wb, sheetname)[1]
            else:    
                xBase = getBasepointDest(wb, sheetname)[0]
                yBase = getBasepointDest(wb, sheetname)[1]            
#            xBase = getBasepointDest(dest, sheetname)[0]
#            yBase = getBasepointDest(dest, sheetname)[1]
            print 'base point= ',xBase,yBase
            writeToDest(wb,sheetname,xBase,yBase,dfReportDataArea)
                                   
stop = timeit.default_timer()
print stop - start

text.insert(INSERT, 'Successfully!\n')
text.insert(INSERT, 'Time Cost: '+str(stop - start)+'s\n')
text.update()
text.see('end') 

root.mainloop()
logfile.write('\n')
logfile.write('Data Transfer Successfully!\n')
logfile.write('Time Cost: '+str(stop - start)+'s')
logfile.close()

#text.insert(INSERT, '耗时'+(stop - start)+'秒'+'\n')
#text.update()
#root.mainloop()            