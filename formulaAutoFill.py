# -*- coding: utf-8 -*-
"""
Created on Tue Sep 19 11:29:35 2017

@author: Maddox.Meng
"""
import random
from openpyxl import load_workbook

excel = r'C:\Workspace\测试\tst-formula.xlsx'.decode('utf-8')
wb = load_workbook(excel)
ws = wb.get_sheet_by_name('Sheet1')

nrows = ws.max_row
ncols = ws.max_column

print ws['C1'].value #=A1+B1

for row in ws['A1:B30']:
    for cell in row:
        cell.value = random.randint(0, 20)

for cell in ws['C']:
#    cell.value = ws['C1'].value.replace(1,cell.rows)
    cell.value = ws['C1'].value.replace('1',str(cell.row))
      
wb.save(excel)        