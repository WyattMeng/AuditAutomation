# -*- coding: utf-8 -*-
"""
Created on Tue Sep 26 15:04:01 2017

@author: Maddox.Meng
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

path = r'C:\Workspace\测试\GET.xlsx'.decode('utf-8')
wb = load_workbook(path)


'''4改3，ws.max_column改成'''            
def getYColumnNo(ytitleRowlist):

    ytitleCollist = []
    for x in ytitleRowlist:
        for y in range(3, ws.max_column): #ytitle都在第5、6列起始。特例：R利率敏感性分析，没有ytitle。
            #1.对yrowslist遍历
            #2.找到非空单元格，记录到list
            if ws[x][y].value != None:
                #print ws[x][y].value, ws[x][y].column
                ytitleCollist.append(ws[x][y].column)
                ytitleCollist = sorted(list(set(ytitleCollist))) #去重、排序
    return ytitleCollist 


sheetnames = []
for sheetname in wb.sheetnames:
    if sheetname != 'Catalog' or sheetname != 'Instraction':
        sheetnames.append(sheetname)


#有<End>才是数据相关的sheetname
dataSheetnames = []        
for sheetname in wb.sheetnames:
    ws = wb.get_sheet_by_name(sheetname)
    for cell in ws['A']:
        if cell.value == '<End>':
            dataSheetnames.append(sheetname)
            break
ret = list(set(wb.sheetnames) ^ set(dataSheetnames)) 

print ret #[u'Qlog', u'Catalog', u'Instraction']      

for sheetname in dataSheetnames:
    ws = wb.get_sheet_by_name(sheetname)
    
    for cell in ws['A']:
        if cell.value == '<End>':
            max_row = cell.row - 1
    
    y=0
    rownumbers_A = []
    for x in range(3, max_row + 1):
        if ws[x][y].value != None:
            rownumbers_A.append(ws[x][y].row)
    
    y=1         
    rownumbers_B = []
    for x in range(3, max_row + 1):
        if ws[x][y].value != None:
            rownumbers_B.append(ws[x][y].row)
    
    if len(rownumbers_A) <= len(rownumbers_B):
        y=1
        rownumbers = rownumbers_B
    else:
        y=0  
        rownumbers = rownumbers_A

    yrows = min(rownumbers) - 2 - 1
    
    ytitleRowlist = []
    for i in range(0, yrows):
        ytitleRowlist.append(2 + i)
        
    #print ytitleRowlist 
    ytitleCollist = getYColumnNo(ytitleRowlist)
    print ytitleCollist

i=0
for sheetname in dataSheetnames:
    ws = wb.get_sheet_by_name(sheetname)
    
    for cell in ws[2]:
        if cell.value == 'Casting': #63
            max_col = column_index_from_string(cell.column) -1 - 1
            #i+=1
            break
        elif cell.value == 'PY':
            max_col = column_index_from_string(cell.column) -1 - 1
            break
            #i+=1
        else:
            max_col = ws.max_column
            i+=1
            
            
            
            

                